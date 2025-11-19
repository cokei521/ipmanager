from flask import Flask, render_template, redirect, url_for, request, flash, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from datetime import datetime
import os
import logging
import ipaddress
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ip_addresses.db'

# 添加请求日志
@app.before_request
def log_request():
    logger.info(f'请求路径: {request.path}')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 配置静态文件和模板文件夹
app.config['TEMPLATES_FOLDER'] = 'templates'

db = SQLAlchemy(app)

# 配置Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录以访问该页面'
login_manager.login_message_category = 'info'

# 角色/权限模型
class Role(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    can_add_ip = db.Column(db.Boolean, default=False)
    can_edit_ip = db.Column(db.Boolean, default=False)
    can_delete_ip = db.Column(db.Boolean, default=False)
    can_view_ip = db.Column(db.Boolean, default=True)
    can_manage_users = db.Column(db.Boolean, default=False)

    def __repr__(self):
        return f'<Role {self.name}>'

# 用户模型
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('role.id'), nullable=False)
    role = db.relationship('Role', backref=db.backref('users', lazy=True))

    def __repr__(self):
        return f'<User {self.username}>'


# 加载用户的回调函数
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# 权限控制装饰器
def require_permission(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not getattr(current_user.role, permission, False):
                flash('您没有权限执行此操作！', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# 查看IP权限装饰器
def can_view_ip(f):
    return require_permission('can_view_ip')(f)

# 添加IP权限装饰器
def can_add_ip(f):
    return require_permission('can_add_ip')(f)

# 编辑IP权限装饰器
def can_edit_ip(f):
    return require_permission('can_edit_ip')(f)

# 删除IP权限装饰器
def can_delete_ip(f):
    return require_permission('can_delete_ip')(f)

# 管理用户权限装饰器
def can_manage_users(f):
    return require_permission('can_manage_users')(f)

# IP地址模型
class IPAddress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    logical_division = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(15), unique=True, nullable=False, index=True)
    subnet_mask = db.Column(db.String(15), nullable=False)
    in_use = db.Column(db.Boolean, nullable=False, default=True)
    device_name = db.Column(db.String(100))
    system = db.Column(db.String(100))
    related_device = db.Column(db.String(100))
    department = db.Column(db.String(100))
    responsible_person = db.Column(db.String(100))
    responsible_person_phone = db.Column(db.String(20))
    vendor_responsible_person = db.Column(db.String(100))
    vendor_responsible_person_phone = db.Column(db.String(20))
    work_order_number = db.Column(db.String(50))
    location = db.Column(db.String(100))
    is_scanned = db.Column(db.Boolean, nullable=False, default=False)
    update_date = db.Column(db.Date, nullable=False, default=datetime.now().date())
    remarks = db.Column(db.Text)

    def __repr__(self):
        return f'<IPAddress {self.ip_address}>'

# 辅助函数：验证IP地址格式
def is_valid_ip(ip):
    parts = ip.split('.')
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(part) <= 255 for part in parts)
    except ValueError:
        return False

# 辅助函数：按IP地址排序
def get_ip_numeric(ip):
    parts = ip.split('.')
    return int(parts[0]) * 256**3 + int(parts[1]) * 256**2 + int(parts[2]) * 256 + int(parts[3])

# Bootstrap现在通过CDN加载，不再需要本地路由

# 登录路由
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash('登录成功！', 'success')
            return redirect(url_for('index'))
        else:
            flash('用户名或密码错误！', 'danger')
    return render_template('login.html')

# 注销路由
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('已成功注销！', 'info')
    return redirect(url_for('login'))

# 用户管理路由
@app.route('/users')
@login_required
@can_manage_users
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

# 添加用户路由
@app.route('/add_user', methods=['GET', 'POST'])
@login_required
@can_manage_users
def add_user():
    roles = Role.query.all()
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        role_id = request.form['role_id']
        
        # 检查用户名是否已存在
        if User.query.filter_by(username=username).first():
            flash('用户名已存在！', 'danger')
            return redirect(url_for('add_user'))
        
        # 检查密码和确认密码是否一致
        if password != confirm_password:
            flash('密码和确认密码不一致！', 'danger')
            return redirect(url_for('add_user'))
        
        # 检查密码长度
        if len(password) < 6:
            flash('密码长度至少为6个字符！', 'danger')
            return redirect(url_for('add_user'))
        
        # 创建用户
        hashed_password = generate_password_hash(password)
        user = User(username=username, password=hashed_password, role_id=role_id)
        
        # 更新角色权限
        role = Role.query.get(role_id)
        role.can_view_ip = 'can_view_ip' in request.form
        role.can_add_ip = 'can_add_ip' in request.form
        role.can_edit_ip = 'can_edit_ip' in request.form
        role.can_delete_ip = 'can_delete_ip' in request.form
        role.can_manage_users = 'can_manage_users' in request.form
        
        db.session.add(user)
        db.session.commit()
        flash('用户添加成功！', 'success')
        return redirect(url_for('users'))
    return render_template('add_user.html', roles=roles)

# 编辑用户路由
@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@can_manage_users
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    roles = Role.query.all()
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role_id = request.form['role_id']
        
        # 检查用户名是否已存在（排除当前用户）
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != user_id:
            flash('用户名已存在！', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        
        # 更新用户名
        user.username = username
        
        # 如果提供了新密码，则更新密码
        if password:
            if len(password) < 6:
                flash('密码长度至少为6个字符！', 'danger')
                return redirect(url_for('edit_user', user_id=user_id))
            if password != request.form['confirm_password']:
                flash('密码和确认密码不一致！', 'danger')
                return redirect(url_for('edit_user', user_id=user_id))
            user.password = generate_password_hash(password)
        
        # 更新角色
        user.role_id = role_id
        
        # 更新角色权限
        role = Role.query.get(role_id)
        role.can_view_ip = 'can_view_ip' in request.form
        role.can_add_ip = 'can_add_ip' in request.form
        role.can_edit_ip = 'can_edit_ip' in request.form
        role.can_delete_ip = 'can_delete_ip' in request.form
        role.can_manage_users = 'can_manage_users' in request.form
        
        db.session.commit()
        flash('用户信息更新成功！', 'success')
        return redirect(url_for('users'))
    return render_template('edit_user.html', user=user, roles=roles)

# 删除用户路由
@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@can_manage_users
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # 不能删除最后一个管理员用户
    admin_users = User.query.join(Role).filter(Role.can_manage_users == True).count()
    if admin_users <= 1 and user.role.can_manage_users:
        flash('不能删除最后一个管理员用户！', 'danger')
        return redirect(url_for('users'))
    
    db.session.delete(user)
    db.session.commit()
    flash('用户删除成功！', 'success')
    return redirect(url_for('users'))

# 测试路由，用于验证Flask应用是否正常工作
@app.route('/test')
def test():
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return f"当前时间: {current_time}，Flask应用正常工作！"

# 主页路由
@app.route('/')
@login_required
@can_view_ip
def index():
    page = request.args.get('page', 1, type=int)
    per_page = 300  # 每页显示300行数据
    
    # 处理in_use筛选参数，将字符串'true'和'false'正确转换为布尔值
    in_use_param = request.args.get('in_use')
    in_use_filter = None
    if in_use_param is not None:
        in_use_filter = in_use_param.lower() == 'true'
    
    # 对于SQLite，我们需要手动拆分IP地址的四个部分并转换为整数进行排序
    # 这样可以确保IP地址按数字顺序正确排序
    ip = IPAddress.ip_address
    parts = [
        db.func.cast(db.func.substr(ip, 1, db.func.instr(ip, '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, db.func.instr(ip, '.') + 1, 
                                 db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, 
                                 db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1, 
                                 db.func.instr(db.func.substr(ip, 
                                                          db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1), '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, 
                                 db.func.instr(ip, '.') + 
                                 db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 
                                 db.func.instr(db.func.substr(ip, 
                                                          db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1), '.') + 1), db.Integer)
    ]
    
    # 根据筛选条件查询IP地址
    query = IPAddress.query
    if in_use_filter is not None:
        query = query.filter_by(in_use=in_use_filter)
    
    # 使用拆分后的IP部分进行排序并分页
    ip_addresses = query.order_by(*parts).paginate(page=page, per_page=per_page)
    
    # IP地址统计
    total_ips = IPAddress.query.count()
    used_ips = IPAddress.query.filter_by(in_use=True).count()
    free_ips = IPAddress.query.filter_by(in_use=False).count()
    
    return render_template('index.html', ip_addresses=ip_addresses, page=page, 
                         total_ips=total_ips, used_ips=used_ips, free_ips=free_ips, in_use_filter=in_use_filter)

# 添加IP地址路由
@app.route('/add', methods=['GET', 'POST'])
@login_required
@can_add_ip
def add_ip():
    if request.method == 'POST':
        logical_division = request.form['logical_division']
        ip_address = request.form['ip_address']
        subnet_mask = request.form['subnet_mask']
        in_use = request.form.get('in_use') == 'on'
        device_name = request.form['device_name']
        system = request.form['system']
        related_device = request.form['related_device']
        department = request.form['department']
        responsible_person = request.form['responsible_person']
        responsible_person_phone = request.form['responsible_person_phone']
        vendor_responsible_person = request.form['vendor_responsible_person']
        vendor_responsible_person_phone = request.form['vendor_responsible_person_phone']
        work_order_number = request.form['work_order_number']
        location = request.form['location']
        is_scanned = request.form.get('is_scanned') == 'on'
        update_date_str = request.form['update_date']
        remarks = request.form['remarks']
        
        # 验证IP地址格式
        if not is_valid_ip(ip_address):
            flash('请输入有效的IP地址格式', 'danger')
            return redirect(url_for('add_ip'))
        
        # 处理子网掩码，支持CIDR格式
        processed_subnet_mask = subnet_mask
        try:
            # 检查是否为CIDR格式（纯数字）
            if subnet_mask.isdigit():
                cidr = int(subnet_mask)
                if 0 <= cidr <= 32:
                    # 使用ipaddress库将CIDR转换为点分十进制
                    network = ipaddress.IPv4Network(f'0.0.0.0/{cidr}', strict=False)
                    processed_subnet_mask = str(network.netmask)
        except Exception as e:
            # 如果转换失败，保持原始输入
            pass
        
        # 转换更新日期
        update_date = datetime.strptime(update_date_str, '%Y-%m-%d').date()
        
        try:
            # 尝试将输入的IP和掩码视为网络地址段
            # 使用处理后的子网掩码
            network_str = f"{ip_address}/{processed_subnet_mask}"
            network = ipaddress.IPv4Network(network_str, strict=False)
            
            # 检查输入的IP地址是否为网络地址（即与网络的第一个地址相同）
            if ip_address == str(network.network_address):
                # 输入的是网络地址，批量添加整个网络的IP
                added_count = 0
                skipped_count = 0
                
                # 开始事务
                for ip in network.hosts():
                    ip_str = str(ip)
                    
                    # 检查IP是否已存在
                    existing_ip = IPAddress.query.filter_by(ip_address=ip_str).first()
                    if existing_ip:
                        skipped_count += 1
                        continue
                    
                    # 创建新的IP地址记录
                    new_ip = IPAddress(
                        logical_division=logical_division,
                        ip_address=ip_str,
                        subnet_mask=processed_subnet_mask,
                        in_use=in_use,
                        device_name=device_name,
                        system=system,
                        related_device=related_device,
                        department=department,
                        responsible_person=responsible_person,
                        responsible_person_phone=responsible_person_phone,
                        vendor_responsible_person=vendor_responsible_person,
                        vendor_responsible_person_phone=vendor_responsible_person_phone,
                        work_order_number=work_order_number,
                        location=location,
                        is_scanned=is_scanned,
                        update_date=update_date,
                        remarks=remarks
                    )
                    db.session.add(new_ip)
                    added_count += 1
                
                # 提交事务
                db.session.commit()
                
                message = f'IP地址段添加成功！成功添加 {added_count} 个IP地址'
                if skipped_count > 0:
                    message += f'，跳过 {skipped_count} 个已存在的IP地址'
                flash(message, 'success')
                return redirect(url_for('index'))
            else:
                # 输入的是单个IP地址，仅添加这一个IP
                # 检查IP地址是否已存在
                existing_ip = IPAddress.query.filter_by(ip_address=ip_address).first()
                if existing_ip:
                    flash('该IP地址已存在', 'danger')
                    return redirect(url_for('add_ip'))
                
                # 创建新的IP地址记录
                new_ip = IPAddress(
                    logical_division=logical_division,
                    ip_address=ip_address,
                    subnet_mask=processed_subnet_mask,
                    in_use=in_use,
                    device_name=device_name,
                    system=system,
                    related_device=related_device,
                    department=department,
                    responsible_person=responsible_person,
                    responsible_person_phone=responsible_person_phone,
                    vendor_responsible_person=vendor_responsible_person,
                    vendor_responsible_person_phone=vendor_responsible_person_phone,
                    work_order_number=work_order_number,
                    location=location,
                    is_scanned=is_scanned,
                    update_date=update_date,
                    remarks=remarks
                )
                
                db.session.add(new_ip)
                db.session.commit()
                flash('IP地址添加成功', 'success')
                return redirect(url_for('index'))
                
        except ValueError as e:
            # IP地址或掩码格式无效
            flash('IP地址或子网掩码格式无效：' + str(e), 'danger')
            return redirect(url_for('add_ip'))
        except Exception as e:
            db.session.rollback()
            flash('添加失败：' + str(e), 'danger')
            return redirect(url_for('add_ip'))
    
    # 逻辑划分选项
    logical_divisions = ['MDCN', 'NMS-CE03&04', 'BS-5GC', 'MW-5GC', 'NFV', 'UMF', 'PC']
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('add.html', logical_divisions=logical_divisions, today=today)

# 编辑IP地址路由
@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@can_edit_ip
def edit_ip(id):
    ip = IPAddress.query.get_or_404(id)
    
    if request.method == 'POST':
        logical_division = request.form['logical_division']
        ip_address = request.form['ip_address']
        subnet_mask = request.form['subnet_mask']
        in_use = request.form.get('in_use') == 'on'
        device_name = request.form['device_name']
        system = request.form['system']
        related_device = request.form['related_device']
        department = request.form['department']
        responsible_person = request.form['responsible_person']
        responsible_person_phone = request.form['responsible_person_phone']
        vendor_responsible_person = request.form['vendor_responsible_person']
        vendor_responsible_person_phone = request.form['vendor_responsible_person_phone']
        work_order_number = request.form['work_order_number']
        location = request.form['location']
        is_scanned = request.form.get('is_scanned') == 'on'
        update_date_str = request.form['update_date']
        remarks = request.form['remarks']
        
        # 验证IP地址格式
        if not is_valid_ip(ip_address):
            flash('请输入有效的IP地址格式', 'danger')
            return redirect(url_for('edit_ip', id=id))
        
        # 检查IP地址是否已被其他记录使用
        existing_ip = IPAddress.query.filter_by(ip_address=ip_address).first()
        if existing_ip and existing_ip.id != id:
            flash('该IP地址已被其他记录使用', 'danger')
            return redirect(url_for('edit_ip', id=id))
        
        # 转换更新日期
        update_date = datetime.strptime(update_date_str, '%Y-%m-%d').date()
        
        # 处理子网掩码，支持CIDR格式
        processed_subnet_mask = subnet_mask
        try:
            # 检查是否为CIDR格式（纯数字）
            if subnet_mask.isdigit():
                cidr = int(subnet_mask)
                if 0 <= cidr <= 32:
                    # 使用ipaddress库将CIDR转换为点分十进制
                    network = ipaddress.IPv4Network(f'0.0.0.0/{cidr}', strict=False)
                    processed_subnet_mask = str(network.netmask)
        except Exception as e:
            # 如果转换失败，保持原始输入
            pass
        
        # 更新IP地址记录
        ip.logical_division = logical_division
        ip.ip_address = ip_address
        ip.subnet_mask = processed_subnet_mask
        ip.in_use = in_use
        ip.device_name = device_name
        ip.system = system
        ip.related_device = related_device
        ip.department = department
        ip.responsible_person = responsible_person
        ip.responsible_person_phone = responsible_person_phone
        ip.vendor_responsible_person = vendor_responsible_person
        ip.vendor_responsible_person_phone = vendor_responsible_person_phone
        ip.work_order_number = work_order_number
        ip.location = location
        ip.is_scanned = is_scanned
        ip.update_date = update_date
        ip.remarks = remarks
        
        try:
            db.session.commit()
            flash('IP地址更新成功', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash('更新失败：' + str(e), 'danger')
            return redirect(url_for('edit_ip', id=id))
    
    # 逻辑划分选项
    logical_divisions = ['MDCN', 'NMS-CE03&04', 'BS-5GC', 'MW-5GC', 'NFV', 'UMF', 'PC']
    update_date_formatted = ip.update_date.strftime('%Y-%m-%d')
    return render_template('edit.html', ip=ip, logical_divisions=logical_divisions, update_date_formatted=update_date_formatted)

# 删除IP地址路由
@app.route('/delete/<int:id>', methods=['POST'])
@login_required
@can_delete_ip
def delete_ip(id):
    ip = IPAddress.query.get_or_404(id)
    
    try:
        db.session.delete(ip)
        db.session.commit()
        flash('IP地址删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash('删除失败：' + str(e), 'danger')
    
    return redirect(url_for('index'))

# 批量删除IP地址路由
@app.route('/batch_delete', methods=['POST'])
@login_required
@can_delete_ip
def batch_delete_ip():
    # 获取表单提交的IP ID列表
    ip_ids = request.form.getlist('ip_ids')
    
    if not ip_ids:
        flash('请选择要删除的IP地址', 'warning')
        return redirect(url_for('index'))
    
    try:
        # 将字符串ID转换为整数
        ip_ids = [int(id) for id in ip_ids]
        
        # 查询所有要删除的IP地址
        ips_to_delete = IPAddress.query.filter(IPAddress.id.in_(ip_ids)).all()
        
        # 删除查询到的IP地址
        for ip in ips_to_delete:
            db.session.delete(ip)
        
        # 提交事务
        db.session.commit()
        flash(f'成功删除 {len(ips_to_delete)} 个IP地址', 'success')
    except ValueError as e:
        db.session.rollback()
        flash('无效的IP ID格式', 'danger')
    except Exception as e:
        db.session.rollback()
        flash('批量删除失败：' + str(e), 'danger')
    
    return redirect(url_for('index'))

# 搜索功能
@app.route('/search', methods=['GET'])
@login_required
@can_view_ip
def search():
    query = request.args.get('query', '')
    page = request.args.get('page', 1, type=int)
    per_page = 300  # 与主页保持一致的分页大小
    
    # 构建查询
    if query:
        ip_query = IPAddress.query.filter(
            (IPAddress.ip_address.contains(query)) |
            (IPAddress.device_name.contains(query)) |
            (IPAddress.responsible_person.contains(query)) |
            (IPAddress.department.contains(query))
        )
    else:
        ip_query = IPAddress.query
    
    # 对于SQLite，使用与主页相同的方式拆分IP地址进行排序
    ip = IPAddress.ip_address
    parts = [
        db.func.cast(db.func.substr(ip, 1, db.func.instr(ip, '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, db.func.instr(ip, '.') + 1, 
                                 db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, 
                                 db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1, 
                                 db.func.instr(db.func.substr(ip, 
                                                          db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1), '.') - 1), db.Integer),
        db.func.cast(db.func.substr(ip, 
                                 db.func.instr(ip, '.') + 
                                 db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 
                                 db.func.instr(db.func.substr(ip, 
                                                          db.func.instr(ip, '.') + db.func.instr(db.func.substr(ip, db.func.instr(ip, '.') + 1), '.') + 1), '.') + 1), db.Integer)
    ]
    
    # 使用拆分后的IP部分进行排序并分页
    ip_addresses = ip_query.order_by(*parts).paginate(page=page, per_page=per_page)
    return render_template('index.html', ip_addresses=ip_addresses, search_query=query, page=page)

# 自动补全API
@app.route('/autocomplete', methods=['GET'])
@login_required
@can_view_ip
def autocomplete():
    field = request.args.get('field', '')
    query = request.args.get('query', '')
    
    if not field or not query:
        return {'results': []}
    
    try:
        # 获取指定字段的所有不重复值，且包含查询词
        results = IPAddress.query.with_entities(getattr(IPAddress, field)).distinct()
        results = results.filter(getattr(IPAddress, field).like(f'%{query}%'))
        results = results.filter(getattr(IPAddress, field) != '')
        results = results.all()
        
        # 转换为列表格式
        results_list = [result[0] for result in results]
        
        return {'results': results_list[:10]}  # 最多返回10个结果
    except Exception as e:
        logger.error(f'自动补全API错误: {e}')
        return {'results': []}

# 导出Excel功能
@app.route('/export_excel', methods=['POST'])
@login_required
@can_view_ip
def export_excel():
    # 获取选中的IP ID列表
    ip_ids = request.form.getlist('ip_ids')
    
    if not ip_ids:
        flash('请选择要导出的IP地址', 'warning')
        return redirect(url_for('index'))
    
    try:
        # 将字符串ID转换为整数
        ip_ids = [int(id) for id in ip_ids]
        
        # 查询所有要导出的IP地址
        ips_to_export = IPAddress.query.filter(IPAddress.id.in_(ip_ids)).all()
        
        # 按IP地址从小到大排序
        ips_to_export = sorted(ips_to_export, key=lambda x: get_ip_numeric(x.ip_address))
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "IP地址列表"
        
        # 设置表头
        headers = ['逻辑划分', 'IP地址', '掩码', '在用', '设备名称', '所属系统', '关联设备', '所属科室', 
                   '责任人', '责任人电话', '厂家责任人', '厂家责任人电话', '工单号', '位置', '是否扫描', 
                   '更新日期', '备注']
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row, ip in enumerate(ips_to_export, 2):
            ws.cell(row=row, column=1, value=ip.logical_division)
            ws.cell(row=row, column=2, value=ip.ip_address)
            ws.cell(row=row, column=3, value=ip.subnet_mask)
            ws.cell(row=row, column=4, value='是' if ip.in_use else '否')
            ws.cell(row=row, column=5, value=ip.device_name or '')
            ws.cell(row=row, column=6, value=ip.system or '')
            ws.cell(row=row, column=7, value=ip.related_device or '')
            ws.cell(row=row, column=8, value=ip.department or '')
            ws.cell(row=row, column=9, value=ip.responsible_person or '')
            ws.cell(row=row, column=10, value=ip.responsible_person_phone or '')
            ws.cell(row=row, column=11, value=ip.vendor_responsible_person or '')
            ws.cell(row=row, column=12, value=ip.vendor_responsible_person_phone or '')
            ws.cell(row=row, column=13, value=ip.work_order_number or '')
            ws.cell(row=row, column=14, value=ip.location or '')
            ws.cell(row=row, column=15, value='是' if ip.is_scanned else '否')
            ws.cell(row=row, column=16, value=ip.update_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=17, value=ip.remarks or '')
        
        # 设置列宽自适应
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # 保存工作簿到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ip_addresses_{current_time}.xlsx'
        
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ValueError as e:
        db.session.rollback()
        flash('无效的IP ID格式', 'danger')
        return redirect(url_for('index'))
    except Exception as e:
        logger.error(f'导出Excel失败: {e}')
        flash('导出Excel失败：' + str(e), 'danger')
        return redirect(url_for('index'))

# 导入Excel功能
@app.route('/import_excel', methods=['POST'])
@login_required
@can_add_ip
def import_excel():
    if 'file' not in request.files:
        flash('请选择要导入的Excel文件', 'warning')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('请选择要导入的Excel文件', 'warning')
        return redirect(url_for('index'))
    
    if not file.filename.endswith('.xlsx'):
        flash('只支持导入.xlsx格式的Excel文件', 'danger')
        return redirect(url_for('index'))
    
    try:
        from openpyxl import load_workbook
        
        # 加载Excel文件
        wb = load_workbook(file)
        ws = wb.active
        
        # 读取表头
        headers = [cell.value for cell in ws[1]]
        
        # 检查表头是否符合要求
        required_headers = ['IP地址', '掩码', '逻辑划分']
        for header in required_headers:
            if header not in headers:
                flash(f'Excel文件缺少必要的列：{header}', 'danger')
                return redirect(url_for('index'))
        
        # 获取列索引
        ip_col = headers.index('IP地址') + 1
        mask_col = headers.index('掩码') + 1
        division_col = headers.index('逻辑划分') + 1
        
        # 其他可选列的索引
        in_use_col = headers.index('在用') + 1 if '在用' in headers else None
        device_name_col = headers.index('设备名称') + 1 if '设备名称' in headers else None
        system_col = headers.index('所属系统') + 1 if '所属系统' in headers else None
        related_device_col = headers.index('关联设备') + 1 if '关联设备' in headers else None
        department_col = headers.index('所属科室') + 1 if '所属科室' in headers else None
        responsible_person_col = headers.index('责任人') + 1 if '责任人' in headers else None
        responsible_person_phone_col = headers.index('责任人电话') + 1 if '责任人电话' in headers else None
        vendor_responsible_person_col = headers.index('厂家责任人') + 1 if '厂家责任人' in headers else None
        vendor_responsible_person_phone_col = headers.index('厂家责任人电话') + 1 if '厂家责任人电话' in headers else None
        work_order_number_col = headers.index('工单号') + 1 if '工单号' in headers else None
        location_col = headers.index('位置') + 1 if '位置' in headers else None
        is_scanned_col = headers.index('是否扫描') + 1 if '是否扫描' in headers else None
        update_date_col = headers.index('更新日期') + 1 if '更新日期' in headers else None
        remarks_col = headers.index('备注') + 1 if '备注' in headers else None
        
        # 处理数据行
        added_count = 0
        updated_count = 0
        skipped_count = 0
        
        for row in ws.iter_rows(min_row=2):
            ip_address = row[ip_col-1].value
            
            # 跳过空行
            if not ip_address:
                continue
            
            # 验证IP地址格式
            if not is_valid_ip(str(ip_address)):
                skipped_count += 1
                continue
            
            # 获取数据
            subnet_mask = row[mask_col-1].value or ''
            logical_division = row[division_col-1].value or ''
            in_use = row[in_use_col-1].value == '是' if in_use_col else True
            device_name = row[device_name_col-1].value or '' if device_name_col else ''
            system = row[system_col-1].value or '' if system_col else ''
            related_device = row[related_device_col-1].value or '' if related_device_col else ''
            department = row[department_col-1].value or '' if department_col else ''
            responsible_person = row[responsible_person_col-1].value or '' if responsible_person_col else ''
            responsible_person_phone = row[responsible_person_phone_col-1].value or '' if responsible_person_phone_col else ''
            vendor_responsible_person = row[vendor_responsible_person_col-1].value or '' if vendor_responsible_person_col else ''
            vendor_responsible_person_phone = row[vendor_responsible_person_phone_col-1].value or '' if vendor_responsible_person_phone_col else ''
            work_order_number = row[work_order_number_col-1].value or '' if work_order_number_col else ''
            location = row[location_col-1].value or '' if location_col else ''
            is_scanned = row[is_scanned_col-1].value == '是' if is_scanned_col else False
            update_date = row[update_date_col-1].value or datetime.now().date() if update_date_col else datetime.now().date()
            remarks = row[remarks_col-1].value or '' if remarks_col else ''
            
            # 处理子网掩码，支持CIDR格式
            processed_subnet_mask = subnet_mask
            try:
                # 检查是否为CIDR格式（纯数字）
                if processed_subnet_mask and str(processed_subnet_mask).isdigit():
                    cidr = int(processed_subnet_mask)
                    if 0 <= cidr <= 32:
                        # 使用ipaddress库将CIDR转换为点分十进制
                        network = ipaddress.IPv4Network(f'0.0.0.0/{cidr}', strict=False)
                        processed_subnet_mask = str(network.netmask)
            except Exception as e:
                # 如果转换失败，保持原始输入
                pass
            
            # 转换更新日期
            if isinstance(update_date, str):
                try:
                    update_date = datetime.strptime(update_date, '%Y-%m-%d').date()
                except:
                    update_date = datetime.now().date()
            
            # 检查IP是否已存在
            existing_ip = IPAddress.query.filter_by(ip_address=ip_address).first()
            
            if existing_ip:
                # 更新现有记录
                existing_ip.logical_division = logical_division
                existing_ip.subnet_mask = processed_subnet_mask
                existing_ip.in_use = in_use
                existing_ip.device_name = device_name
                existing_ip.system = system
                existing_ip.related_device = related_device
                existing_ip.department = department
                existing_ip.responsible_person = responsible_person
                existing_ip.responsible_person_phone = responsible_person_phone
                existing_ip.vendor_responsible_person = vendor_responsible_person
                existing_ip.vendor_responsible_person_phone = vendor_responsible_person_phone
                existing_ip.work_order_number = work_order_number
                existing_ip.location = location
                existing_ip.is_scanned = is_scanned
                existing_ip.update_date = update_date
                existing_ip.remarks = remarks
                updated_count += 1
            else:
                # 创建新记录
                new_ip = IPAddress(
                    logical_division=logical_division,
                    ip_address=ip_address,
                    subnet_mask=processed_subnet_mask,
                    in_use=in_use,
                    device_name=device_name,
                    system=system,
                    related_device=related_device,
                    department=department,
                    responsible_person=responsible_person,
                    responsible_person_phone=responsible_person_phone,
                    vendor_responsible_person=vendor_responsible_person,
                    vendor_responsible_person_phone=vendor_responsible_person_phone,
                    work_order_number=work_order_number,
                    location=location,
                    is_scanned=is_scanned,
                    update_date=update_date,
                    remarks=remarks
                )
                db.session.add(new_ip)
                added_count += 1
        
        # 提交事务
        db.session.commit()
        
        message = f'Excel导入成功！'
        if added_count > 0:
            message += f' 新增 {added_count} 个IP地址'
        if updated_count > 0:
            message += f'，更新 {updated_count} 个IP地址'
        if skipped_count > 0:
            message += f'，跳过 {skipped_count} 行（格式错误或空行）'
        
        flash(message, 'success')
        return redirect(url_for('index'))
    except Exception as e:
        db.session.rollback()
        logger.error(f'导入Excel失败: {e}')
        flash('导入Excel失败：' + str(e), 'danger')
        return redirect(url_for('index'))

# 初始化数据库
with app.app_context():
    db.create_all()
    
    # 创建初始管理员角色
    admin_role = Role.query.filter_by(name='admin').first()
    if not admin_role:
        admin_role = Role(
            name='admin',
            can_view_ip=True,
            can_add_ip=True,
            can_edit_ip=True,
            can_delete_ip=True,
            can_manage_users=True
        )
        db.session.add(admin_role)
        db.session.commit()
    
    # 创建初始用户角色（仅查看权限）
    user_role = Role.query.filter_by(name='user').first()
    if not user_role:
        user_role = Role(
            name='user',
            can_view_ip=True,
            can_add_ip=False,
            can_edit_ip=False,
            can_delete_ip=False,
            can_manage_users=False
        )
        db.session.add(user_role)
        db.session.commit()
    
    # 创建初始管理员用户
    admin_user = User.query.filter_by(username='admin123').first()
    if not admin_user:
        admin_user = User(
            username='admin123',
            password=generate_password_hash('admin123'),
            role=admin_role
        )
        db.session.add(admin_user)
        db.session.commit()
        print('初始管理员用户创建成功：用户名=admin123，密码=admin123')

# 已使用Flask的static文件夹提供静态文件

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')